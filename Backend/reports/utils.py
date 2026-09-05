from django.http import HttpResponse

from openpyxl import Workbook

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)

from reportlab.lib.styles import getSampleStyleSheet

from reportlab.lib.pagesizes import landscape, letter

from datetime import datetime, date

import io
import os
from django.conf import settings




def clean_value(value):
    """
    Convert unsupported Excel/PDF values
    """


    # List or dictionary fields
    if isinstance(value, (list, dict)):

        return str(value)



    # Remove timezone from datetime
    if isinstance(value, datetime):

        if value.tzinfo:

            value = value.replace(
                tzinfo=None
            )

        return value



    # Date field
    if isinstance(value, date):

        return value



    return value





# =====================================================
# EXCEL EXPORT
# =====================================================


def generate_excel(data, filename):


    workbook = Workbook()


    # Remove default sheet

    sheet = workbook.active

    sheet.title = "Overall Report"



    row_number = 1



    for section, records in data.items():


        # Section title

        sheet.cell(
            row=row_number,
            column=1,
            value=section.upper()
        )


        row_number += 2



        # Empty data

        if not records:


            sheet.cell(
                row=row_number,
                column=1,
                value="No Data Available"
            )


            row_number += 3

            continue




        # Headers

        headers = list(
            records[0].keys()
        )



        for column, header in enumerate(
            headers,
            start=1
        ):

            sheet.cell(
                row=row_number,
                column=column,
                value=header
            )



        row_number += 1




        # Data rows

        for record in records:



            for column, value in enumerate(
                record.values(),
                start=1
            ):


                sheet.cell(
                    row=row_number,
                    column=column,
                    value=clean_value(value)
                )



            row_number += 1



        row_number += 3




    response = HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



    response[
        "Content-Disposition"
    ] = (
        f'attachment; filename="{filename}.xlsx"'
    )



    workbook.save(response)



    return response







# =====================================================
# PDF EXPORT
# =====================================================


def generate_pdf(data, filename):


    buffer = io.BytesIO()



    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter)
    )



    elements = []



    styles = getSampleStyleSheet()





    for section, records in data.items():



        elements.append(
            Paragraph(
                section.upper(),
                styles["Heading2"]
            )
        )



        elements.append(
            Spacer(
                1,
                12
            )
        )



        if not records:


            elements.append(
                Paragraph(
                    "No Data Available",
                    styles["Normal"]
                )
            )


            elements.append(
                Spacer(
                    1,
                    20
                )
            )


            continue





        headers = list(
            records[0].keys()
        )



        table_data = []



        table_data.append(
            headers
        )





        for record in records:


            row = []



            for value in record.values():


                value = clean_value(value)


                row.append(
                    str(value)
                )



            table_data.append(
                row
            )





        table = Table(
            table_data,
            repeatRows=1
        )



        table.setStyle(
            TableStyle(
                [

                    (
                        "GRID",
                        (0,0),
                        (-1,-1),
                        0.5,
                        None
                    ),


                    (
                        "VALIGN",
                        (0,0),
                        (-1,-1),
                        "TOP"
                    )

                ]
            )
        )



        elements.append(
            table
        )



        elements.append(
            Spacer(
                1,
                25
            )
        )





    document.build(
        elements
    )



    pdf = buffer.getvalue()

    buffer.close()

    # Save a copy to the central pdf_storage directory
    pdf_dir = os.path.join(settings.BASE_DIR, 'media', 'pdfs')
    os.makedirs(pdf_dir, exist_ok=True)
    pdf_path = os.path.join(pdf_dir, f"{filename}.pdf")
    with open(pdf_path, 'wb') as f:
        f.write(pdf)

    response = HttpResponse(
        pdf,
        content_type="application/pdf"
    )



    response[
        "Content-Disposition"
    ] = (
        f'attachment; filename="{filename}.pdf"'
    )

    return response
